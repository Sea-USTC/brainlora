import argparse
import os
import yaml as yaml
import numpy as np
import random
import time
import datetime
import json
from pathlib import Path
import openpyxl

import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.utils.data import DataLoader
from sklearn.metrics import roc_auc_score,precision_recall_curve,accuracy_score,confusion_matrix,average_precision_score

# from models.model_MedKLIP import MedKLIP
from models.model_MedKLIP_before_fuse import MedKLIP as MedKLIP
from models.model_MedKLIP_14class_before_fuse import MedKLIP as MedKLIP_14
from models.VIT_image_encoder.VIT_ie import VIT_ie
from dataset.dataset_v1 import MedKLIP_Dataset
from models.tokenization_bert import BertTokenizer
from transformers import AutoModel
from models.imageEncoder import ModelRes, ModelDense
from models.before_fuse import *
from models.SimpleMLP import SimpleMLP

import tqdm
import os
from utils import plot_auc_pr, check_pred

from einops import rearrange
# chexray14_cls = [ 'atelectasis', 'cardiomegaly', 'effusion', 'infiltrate', 'mass', 'nodule', 'pneumonia',
#                 'pneumothorax', 'consolidation', 'edema', 'emphysema', 'tail_abnorm_obs', 'thicken', 'hernia']  #Fibrosis seldom appears in MIMIC_CXR and is divided into the 'tail_abnorm_obs' entitiy.  

# original_class = [
#             'normal', 'clear', 'sharp', 'sharply', 'unremarkable', 'intact', 'stable', 'free',
#             'effusion', 'opacity', 'pneumothorax', 'edema', 'atelectasis', 'tube', 'consolidation', 'process', 'abnormality', 'enlarge', 'tip', 'low',
#             'pneumonia', 'line', 'congestion', 'catheter', 'cardiomegaly', 'fracture', 'air', 'tortuous', 'lead', 'disease', 'calcification', 'prominence',
#             'device', 'engorgement', 'picc', 'clip', 'elevation', 'expand', 'nodule', 'wire', 'fluid', 'degenerative', 'pacemaker', 'thicken', 'marking', 'scar',
#             'hyperinflate', 'blunt', 'loss', 'widen', 'collapse', 'density', 'emphysema', 'aerate', 'mass', 'crowd', 'infiltrate', 'obscure', 'deformity', 'hernia',
#             'drainage', 'distention', 'shift', 'stent', 'pressure', 'lesion', 'finding', 'borderline', 'hardware', 'dilation', 'chf', 'redistribution', 'aspiration',
#             'tail_abnorm_obs', 'excluded_obs'
#         ]
def seed_torch(seed):
    os.environ['PYTHONHASHSEED'] = str(seed) 
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.cuda.manual_seed_all(seed) # if you are using multi-GPU.
    torch.backends.cudnn.benchmark = False
    torch.backends.cudnn.deterministic = True



def _get_bert_basemodel(bert_model_name):
    try:
        model = AutoModel.from_pretrained(bert_model_name)#, return_dict=True)
        print("text feature extractor:", bert_model_name)
    except:
        raise ("Invalid model name. Check the config file and pass a BERT model from transformers lybrary")

    for param in model.parameters():
        param.requires_grad = False

    return model

def get_text_features(model,text_list,tokenizer,device,max_length):
    # text_token =  tokenizer(list(text_list),add_special_tokens=True,max_length=max_length,pad_to_max_length=True,return_tensors='pt').to(device=device)
    target_tokenizer = tokenizer(list(text_list), padding='max_length', truncation=True, max_length=max_length,return_tensors="pt").to(device)
    # text_features = model.encode_text(text_token)
    text_features = model(input_ids = target_tokenizer['input_ids'],attention_mask = target_tokenizer['attention_mask'])#(**encoded_inputs)
    text_features = text_features.last_hidden_state[:,0,:]
    # text_features = F.normalize(text_features, dim=-1)
    return text_features

def compute_AUCs(gt, pred, n_class):
    """Computes Area Under the Curve (AUC) from prediction scores.
    Args:
        gt: Pytorch tensor on GPU, shape = [n_samples, n_classes]
          true binary labels.
        pred: Pytorch tensor on GPU, shape = [n_samples, n_classes]
          can either be probability estimates of the positive class,
          confidence values, or binary decisions.
    Returns:
        List of AUROCs of all classes.
    """
    target_class=json.load(open(config['disease_order'],'r'))
    AUROCs = []
    gt_np = gt.cpu().numpy()
    pred_np = pred.cpu().numpy()
    print("n_classes",n_class)
    for i in range(n_class):
        # print(target_class[i])
        cur_gt = gt_np[:,i]
        cur_pred = pred_np[:,i]
        # print("before",set(cur_gt))
        if strict_test:
            Mask = (( cur_gt!= -1) & ( cur_gt != 2)).squeeze()
            cur_gt = cur_gt[Mask]
            cur_pred = cur_pred[Mask]
        else:
            Mask = (cur_gt == -1).squeeze()
            cur_gt[Mask] = cur_pred[Mask]
        # if 1 not in cur_gt:
        #     print("1 not in gt")
        # elif 0 not in cur_gt:
        #     print("0 not in gt")
        # print("after",set(cur_gt))
        
        AUROCs.append(roc_auc_score(cur_gt, cur_pred))
    return AUROCs

def get_tokenizer(tokenizer,target_text):
    
    target_tokenizer = tokenizer(list(target_text), padding='max_length', truncation=True, max_length= 64, return_tensors="pt")
    
    return target_tokenizer

def test(args,config):

    device=torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print("Total CUDA devices: ", torch.cuda.device_count()) 
    torch.set_default_tensor_type('torch.FloatTensor')

    file_key = args.mode+'_file'
    test_dataset =  MedKLIP_Dataset(config[file_key],config['label_file']) 
    test_dataloader = DataLoader(
            test_dataset,
            batch_size=config['test_batch_size'],
            num_workers=4,
            pin_memory=True,
            sampler=None,
            shuffle=True,
            collate_fn=None,
            drop_last=False,
        ) 
                 
    
    print("Creating book")
    all_target_class = json.load(open(config['disease_order'],'r'))
    target_class = all_target_class.copy()
    # target_class=json.load(open(config['disease_order'],'r'))
    if "exclude_class" in config and config["exclude_class"]:
        keep_class_dim = [target_class.index(i) for i in target_class if i not in config["exclude_classes"] ]
        target_class = [target_class[i] for i in keep_class_dim]
        keep_class_dim = [all_target_class.index(i) for i in all_target_class if i not in config["exclude_classes"] ]
        all_target_class = [target_class[i] for i in keep_class_dim]
    json_book = json.load(open(config['disease_book'],'r'))
    json_order=json.load(open(config['disease_order'],'r'))
    disease_book = [json_book[i] for i in json_order]
    # tokenizer = BertTokenizer.from_pretrained(config['text_encoder'])
    # disease_book_tokenizer = get_tokenizer(tokenizer,disease_book).to(device)

    tokenizer = BertTokenizer.from_pretrained(config['text_encoder'])
    text_encoder = _get_bert_basemodel(config['text_encoder']).to(device)
    text_features = get_text_features(text_encoder,disease_book,tokenizer,device,max_length=256)

    if config['model_type']== 'resnet':
        image_encoder =ModelRes(config).to(device)
    elif config['model_type'] == 'densenet':
        image_encoder = ModelDense(config).to(device)
    elif config['model_type'] == 'VIT':
        image_encoder = VIT_ie(config).to(device)

    fuseModule = beforeFuse(config).to(device) # before fusion
    
    print("Creating model")
    if config['seperate_classifier']:
        print("Medklip_14")
        model = MedKLIP_14(config)
    else:
        print("medklip")
        model = MedKLIP(config)

    model = nn.DataParallel(model, device_ids = [i for i in range(torch.cuda.device_count())])
    model = model.to(device)

    print('Load model from checkpoint:',args.model_path)
    checkpoint = torch.load(args.model_path,map_location='cpu') 
    state_dict = checkpoint['model']          
    model.load_state_dict(state_dict)

    # for name, param in model.named_parameters():
    #     if "classifier" in name:
    #         print("init",name,param.shape)
    #         h = rearrange(param,'c d l -> c (d l)')
    #         print(torch.sum(h,dim=1))
    
    # image_encoder.load_state_dict(checkpoint['image_encoder'])
    image_encoder = nn.DataParallel(image_encoder, device_ids = [i for i in range(torch.cuda.device_count())])
    image_encoder = image_encoder.to(device)
    net_dict = image_encoder.state_dict()
    pretrain_dict = {}
    for k, v in checkpoint['image_encoder'].items():
        new_k = k
        if "conv" in k:
            conv_idx = k.index("conv")+6
            new_k = f"{k[:conv_idx]}{k[conv_idx+5:]}"
        if "layer" in k:
            if 'res_features' in k:
                while "layers" in new_k:
                    lyr_idx = new_k.index("layers")
                    new_k = f"{new_k[:lyr_idx]}{new_k[lyr_idx+7:]}"
            else:
                lyr_idx = k.index("layer")+7
                new_k = f"{k[:lyr_idx]}{k[lyr_idx+7:]}"
        if new_k in net_dict.keys():
            pretrain_dict[new_k]=v
    # print(pretrain_dict.keys())
    # exit(0)
    net_dict.update(pretrain_dict) 
    image_encoder.load_state_dict(net_dict)

    # fuseModule = nn.DataParallel(fuseModule, device_ids = [i for i in range(torch.cuda.device_count())])
    fuseModule.load_state_dict(checkpoint['fuseModule'])
    # fuseModule = fuseModule.to(device)

    # initialize the ground truth and output tensor
    gt = torch.FloatTensor()
    pred = torch.FloatTensor()
    fids = []

    # class_p = json.load(open(config['class_p'],'r'))
    # class_p = torch.tensor(config["la_alpha"])*torch.log(torch.tensor([[class_p[i][0]/class_p[i][1]] for i in target_class]))
    
    print("Start testing")
    model.eval()

    pred_name = args.model_path.split('/')[-1].split('.')[0]
    pred_name = '_'.join([args.model_path.split('/')[-2], pred_name])
    print("pred_name",pred_name)
    rootdir = "/root/Test/preds"+pred_name+main_name+"bswgt/"

    # if os.path.exists(rootdir+"pred_"+args.mode+".npy"):
    if True:
    #     print(rootdir+"pred_"+args.mode+".npy")
    #     pred = np.load(rootdir+"pred_"+args.mode+".npy")
    #     gt = np.load(rootdir+"gt_"+args.mode+".npy")
    #     fids = np.load(rootdir+"fids_"+args.mode+".npy")
    #     gt = torch.tensor(gt)
    #     pred = torch.FloatTensor(pred)
    #     if len(gt) != len(pred):
    #         gt = gt[:len(pred)]
    #     print(gt.shape)
    #     print(pred.shape)
    # else:
        for i, sample in enumerate(test_dataloader):
            images = sample['image']  # [(b,x,y,z),(b,x,y,z)]
            # labels = sample['label'].to(device)
            label = sample['label'][:,:].float()

            if "exclude_class" in config and config["exclude_class"]:
                label = label[:,keep_class_dim]

            B = label.shape[0]

            # if B<6:
            #     continue
            gt = torch.cat((gt, label), 0)
            cur_text_features = text_features.unsqueeze(0).repeat(B,1,1)

            image_features = [] # image_features 4 b n d, image_features_pool 4 b d
            image_features_pool = []
            for idx,cur_image in enumerate(images):
                if config['4_image_encoder']:
                    cur_image_encoder = image_encoder[idx]
                    image_feature,image_feature_pool = cur_image_encoder(cur_image)
                else:
                    image_feature,image_feature_pool = image_encoder(cur_image) 
                image_features.append(image_feature)
                image_features_pool.append(image_feature_pool)
            # before fuse
            fuse_image_feature,_ = fuseModule(image_features)
            #input_image = image.to(device,non_blocking=True)  
            with torch.no_grad():
                pred_class = model(fuse_image_feature,cur_text_features, return_ws=False) #batch_size,num_class,1
                # pred_class = F.softmax(pred_class.reshape(-1,2)).reshape(-1,len(target_class),2)
                # pred_class = pred_class[:,:-1,1]
                # pred_class = pred_class[:,:,1]
                if config["la"]:
                    logits = pred_class.reshape(-1, 1)
                    # cur_class_p = class_p.unsqueeze(0).repeat(B,1,1)
                    # cur_class_p = cur_class_p.reshape(-1,cur_class_p.shape[-1])
                    # cur_class_p = cur_class_p.to(logits.device)
                    # logits = logits - cur_class_p
                    pred_class = torch.sigmoid(logits).reshape(-1,len(all_target_class))
                else:
                    pred_class = torch.sigmoid(pred_class.reshape(-1,1)).reshape(-1,len(all_target_class))

                pred = torch.cat((pred, pred_class.detach().cpu()), 0)
                fids += sample["fid"]
                print("fids",pred.shape,gt.shape)
                # np.save("/home/ps/leijiayu/CODE/MedKLIP/Test_Set_Classification_6thself/preds/"+pred_name+"_fids.npy",np.array(fids))
        
        # os.makedirs(rootdir,exist_ok=True)
        # np.save(rootdir+"pred_"+args.mode+".npy",pred.numpy())
        # np.save(rootdir+"gt_"+args.mode+".npy",gt.numpy())
        # np.save(rootdir+"fids_"+args.mode+".npy",np.array(fids))

    # print("pred.shape",pred)
    # AUROCs = compute_AUCs(gt, pred,len(target_class))
    AUROCs=[]
    max_f1s = []
    accs = []
    precisions=[]
    recalls=[]
    tns,fps,fns,tps = [],[],[],[]
    threshs= []
    aps = []
    #for i in range(len(target_class)-1):
    for i in range(len(target_class)):   
        gt_np = gt[:, i].numpy()
        pred_np = pred[:, i].numpy()
        Mask = (( gt_np!= -1) & ( gt_np != 2)).squeeze()
        cur_gt = gt_np[Mask]
        cur_pred = pred_np[Mask]
        precision, recall, thresholds = precision_recall_curve(cur_gt, cur_pred)
        numerator = 2 * recall * precision # dot multiply for list
        denom = recall + precision
        f1_scores = np.divide(numerator, denom, out=np.zeros_like(denom), where=(denom!=0))
        max_f1 = np.max(f1_scores)
        max_f1_thresh = thresholds[np.argmax(f1_scores)]

        if strict_test:
            AUROCs.append(roc_auc_score(cur_gt, cur_pred))
            aps.append(average_precision_score(cur_gt,cur_pred))
            threshs.append(max_f1_thresh)
            precisions.append(precision[np.argmax(f1_scores)])
            recalls.append(recall[np.argmax(f1_scores)])
            max_f1s.append(max_f1)
            accs.append(accuracy_score(cur_gt, cur_pred>max_f1_thresh))
            pred_label = cur_pred >= max_f1_thresh
            tn,fp,fn,tp = confusion_matrix(cur_gt, pred_label).ravel()
            tns.append(tn)
            fps.append(fp)
            fns.append(fn)
            tps.append(tp)
            plot_auc_pr(cur_gt, cur_pred, target_class[i], rootdir, args.mode)
        else:
            cur_gt = gt_np
            cur_pred = pred_np
            pred_label = cur_pred >= max_f1_thresh
            Mask = (cur_gt == -1).squeeze()
            cur_gt[Mask] = pred_label[Mask]
            precision, recall, thresholds = precision_recall_curve(cur_gt, cur_pred)
            numerator = 2 * recall * precision # dot multiply for list
            denom = recall + precision
            f1_scores = np.divide(numerator, denom, out=np.zeros_like(denom), where=(denom!=0))
            max_f1 = np.max(f1_scores)
            max_f1_thresh = thresholds[np.argmax(f1_scores)]
            aps.append(average_precision_score(cur_gt,cur_pred))
            threshs.append(max_f1_thresh)
            precisions.append(precision[np.argmax(f1_scores)])
            recalls.append(recall[np.argmax(f1_scores)])
            max_f1s.append(max_f1)
            accs.append(accuracy_score(cur_gt, cur_pred>max_f1_thresh))
            pred_label = cur_pred >= max_f1_thresh
            tn,fp,fn,tp = confusion_matrix(cur_gt, pred_label).ravel()
            tns.append(tn)
            fps.append(fp)
            fns.append(fn)
            tps.append(tp)
            plot_auc_pr(cur_gt, cur_pred, target_class[i], rootdir, args.mode)
            AUROCs.append(roc_auc_score(cur_gt, cur_pred))
    
    check_pred(target_class,np.array(fids),threshs,pred,gt,strict_test,rootdir+"result_"+args.mode+".xlsx")

    print('The average AUROC is {AUROC_avg:.4f}'.format(AUROC_avg=np.array(AUROCs[:]).mean()))
    for i in range(len(target_class)):
        print('The AUROC of {} is {}'.format(target_class[i], AUROCs[i]))

    print('The average f1 is {AUROC_avg:.4f}'.format(AUROC_avg=np.array(max_f1s[:]).mean()))
    for i in range(len(target_class)):
        print('The f1 of {} is {}'.format(target_class[i], max_f1s[i]))
    
    print('The average ap is {AUROC_avg:.4f}'.format(AUROC_avg=np.array(aps[:]).mean()))
    for i in range(len(target_class)):
        print('The ap of {} is {}'.format(target_class[i], aps[i]))
    
    print('The average acc is {AUROC_avg:.4f}'.format(AUROC_avg=np.array(accs[:]).mean()))
    for i in range(len(target_class)):
        print('The acc of {} is {}'.format(target_class[i], accs[i]))

    # for i in range(len(target_class)-1):
    print('The average recall is {AUROC_avg:.4f}'.format(AUROC_avg=np.array(recalls[:]).mean()))
    for i in range(len(target_class)):
        print('The recall of {} is {}'.format(target_class[i], recalls[i]))
    print('The average precision is {AUROC_avg:.4f}'.format(AUROC_avg=np.array(precisions[:]).mean()))
    for i in range(len(target_class)):
        print('The precision of {} is {}'.format(target_class[i], precisions[i]))
    #for i in range(len(target_class)-1):
    print('The average thresh is {AUROC_avg:.4f}'.format(AUROC_avg=np.array(threshs[:]).mean()))
    for i in range(len(target_class)):
        print('The thresh of {} is {}'.format(target_class[i], threshs[i]))
        
    # f1_avg = np.array(max_f1s).mean()    
    # acc_avg = np.array(accs).mean()
    # print("all ACC", accs)
    # print('The average f1 is {F1_avg:.4f}'.format(F1_avg=f1_avg))
    # print('The average ACC is {ACC_avg:.4f}'.format(ACC_avg=acc_avg))
    
    # write into excel, AUROCs, max_f1s, accs, aps
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    xlsx_folder_path = '/root/Test/xlsx_files_base_lwgt_npre/'
    xlsx_file_path = xlsx_folder_path + 'baseline_'+ pred_name+'.xlsx'
    if not os.path.exists(xlsx_folder_path):
        os.makedirs(xlsx_folder_path)

    for i, value in enumerate(['auc', 'f1', 'acc', 'ap'], start=1):
        sheet.cell(row=i+1, column=1, value=value)
    for i, value in enumerate(target_class, start=1):
        sheet.cell(row=1, column=i+2, value=value)
    for i, value in enumerate(AUROCs, start=1):
        sheet.cell(row=2, column=i+2, value=value)
    for i, value in enumerate(max_f1s, start=1):
        sheet.cell(row=3, column=i+2, value=value)
    for i, value in enumerate(accs, start=1):
        sheet.cell(row=4, column=i+2, value=value)
    for i, value in enumerate(aps, start=1):
        sheet.cell(row=5, column=i+2, value=value)
    for i, value in enumerate(['aAUC','af1','aACC','aAP'], start=1):
        sheet.cell(row=6, column=i+2, value=value)
    for i, value in enumerate([np.array(AUROCs[:]).mean(), np.array(max_f1s[:]).mean(), np.array(accs[:]).mean(), np.array(aps[:]).mean()], start=1):
        sheet.cell(row=7, column=i+2, value=value)
    workbook.save(xlsx_file_path)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--config', default='/remote-home/mengxichen/UniBrain-lora/Test/configs/config_fifteen.yaml')
    parser.add_argument('--model_path', default='/remote-home/mengxichen/UniBrain-lora/Pretrain/output_fifteen/output_lora1/best_val.pth')
    parser.add_argument('--device', default='cuda')
    parser.add_argument('--gpu', type=str,default='1', help='gpu')
    parser.add_argument('--mode', type=str, default='test')
    parser.add_argument('--seed', type=int,default=2024, help='gpu')
    args = parser.parse_args()

    config = yaml.load(open(args.config, 'r'), Loader=yaml.Loader)

    os.environ["CUDA_VISIBLE_DEVICES"] = args.gpu
    if args.gpu != '-1':
        torch.cuda.current_device()
        torch.cuda._initialized = True
    
    strict_test = True
    main_name = "_nostrict" if not strict_test else ""

    seed_torch(args.seed)  
    test(args, config)